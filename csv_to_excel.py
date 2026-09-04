#!/usr/bin/env python3
"""
Flight Log CSV to Excel Converter
Converts Aphelion GCS CSV exports to formatted Excel workbooks

Usage:
    python csv_to_excel.py ROCKET_FlightLog_*.csv
    python csv_to_excel.py CANSAT_FlightLog_*.csv
"""

import sys
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

def create_excel_from_csv(csv_path):
    """Convert a flight log CSV to a formatted Excel workbook."""
    
    csv_path = Path(csv_path)
    if not csv_path.exists():
        print(f"Error: File {csv_path} not found")
        return False
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Log"
    
    # Read CSV and write to Excel
    with open(csv_path, 'r') as f:
        reader = csv.reader(f)
        
        # Headers
        headers = next(reader)
        
        # Write headers with formatting
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="right", vertical="center")
        
        for row_num, row in enumerate(reader, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Try to convert to float for numeric columns
                try:
                    if col_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:  # All numeric columns
                        cell.value = float(value)
                        if col_num in [1, 2]:  # Time columns
                            cell.number_format = '0.000'
                        elif col_num in [5, 6]:  # Altitude
                            cell.number_format = '0.000'
                        elif col_num in [8, 9]:  # Pressure, Temp
                            cell.number_format = '0.00'
                        elif col_num in [10, 11, 12]:  # Angles
                            cell.number_format = '0.00'
                        elif col_num in [3, 4]:  # Lat/Lon
                            cell.number_format = '0.000000'
                        elif col_num == 7:  # Satellites
                            cell.number_format = '0'
                    else:
                        cell.value = value
                except:
                    cell.value = value
                
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary sheet
    summary = wb.create_sheet("Summary")
    
    # Summary statistics
    summary['A1'] = "Flight Log Summary"
    summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    flight_log = ws
    data_rows = flight_log.max_row - 1
    
    summary['A3'] = "Telemetry Packets:"
    summary['B3'] = data_rows
    summary['B3'].number_format = '0'
    
    # Find max altitude
    if data_rows > 0:
        gps_alt_col = None
        for col_num, header in enumerate(headers, 1):
            if 'GPS Alt' in header:
                gps_alt_col = col_num
                break
        
        if gps_alt_col:
            max_alt = max([
                float(flight_log.cell(row=r, column=gps_alt_col).value or 0)
                for r in range(2, flight_log.max_row + 1)
            ])
            summary['A4'] = "Max GPS Altitude (m):"
            summary['B4'] = max_alt
            summary['B4'].number_format = '0.000'
    
    # Flight duration
    if data_rows > 0:
        time_col = None
        for col_num, header in enumerate(headers, 1):
            if 'Time (s)' in header:
                time_col = col_num
                break
        
        if time_col:
            start_time = float(flight_log.cell(row=2, column=time_col).value or 0)
            end_time = float(flight_log.cell(row=flight_log.max_row, column=time_col).value or 0)
            duration = end_time - start_time
            summary['A5'] = "Flight Duration (s):"
            summary['B5'] = duration
            summary['B5'].number_format = '0.000'
    
    # Auto-adjust summary columns
    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 20
    
    # Save Excel file
    excel_path = csv_path.with_suffix('.xlsx')
    wb.save(str(excel_path))
    
    print(f"✓ Converted: {csv_path.name} → {excel_path.name}")
    print(f"  Packets: {data_rows}")
    print(f"  Output: {excel_path}")
    
    return True

def main():
    if len(sys.argv) < 2:
        print("Flight Log CSV to Excel Converter")
        print("Usage: python csv_to_excel.py <csv_file> [csv_file2] ...")
        print("\nExample:")
        print("  python csv_to_excel.py ROCKET_FlightLog_2024-12-15T14-30-45.csv")
        sys.exit(1)
    
    success_count = 0
    for csv_file in sys.argv[1:]:
        if create_excel_from_csv(csv_file):
            success_count += 1
    
    print(f"\nConverted {success_count}/{len(sys.argv)-1} files successfully")

if __name__ == '__main__':
    main()
